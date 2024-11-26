from openpyxl import *


class Excel:
	def __init__(self):
		pass

	def get_list_name(self, link):
		name_list = []
		try:
			workbook = load_workbook(link)
			sheet = workbook.active
			first_column_data = [str(sheet.cell(row=i, column=1).value) for i in range(1, sheet.max_row + 1)]

			[name_list.append(item) for item in first_column_data]
			return name_list
		except Exception as ex:
			print(f"Ошибка при чтении файла: {ex}")

